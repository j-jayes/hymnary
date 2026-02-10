"""
Post-processing pipeline: classify scraped tune results using an LLM.

Reads the full hymn_tune_index.json from the scraping phase, sends each hymn
through the OpenAI classifier (3 runs with majority vote), and writes:

  1. data/processed/classifications.json  — full LLM reasoning log
  2. data/processed/hymn_tune_index_filtered.json — only relevant tunes kept
  3. data/processed/summary_filtered.csv — flat CSV with is_relevant column

Usage:
    uv run python -m src.filter_pipeline                  # classify all
    uv run python -m src.filter_pipeline --limit 5        # first 5 only
    uv run python -m src.filter_pipeline --reset          # clear checkpoint
    uv run python -m src.filter_pipeline --model gpt-5.2  # specify model
"""

from __future__ import annotations

import argparse
import sys
import traceback
import re
from pathlib import Path
from typing import Any

import pandas as pd

from src.classifier import classify_hymn, get_client
from src.config import INTERIM_DIR, PROCESSED_DIR, HYMN_TUNE_INDEX_PATH, RAW_DIR
from src.utils import (
    get_logger,
    load_checkpoint,
    read_json,
    save_checkpoint,
    write_json,
)

log = get_logger("filter")

# ──────────────────────────────────────────────
# Output paths
# ──────────────────────────────────────────────

CLASSIFICATIONS_PATH = PROCESSED_DIR / "classifications.json"
FILTERED_INDEX_PATH = PROCESSED_DIR / "hymn_tune_index_filtered.json"
FILTERED_CSV_PATH = PROCESSED_DIR / "summary_filtered.csv"
FILTERED_EXCEL_PATH = PROCESSED_DIR / "hymn_tune_data.xlsx"
FILTER_CHECKPOINT_PATH = INTERIM_DIR / "filter_checkpoint.json"
MP_CSV_PATH = RAW_DIR / "MP HymnTuneNames.csv"


def normalize_title(title: Any) -> str:
    """Normalize string for matching: uppercase, strip non-alphanumeric."""
    if not isinstance(title, str):
        return ""
    return re.sub(r'[^A-Z0-9]', '', title.upper())


# ──────────────────────────────────────────────
# Build filtered outputs
# ──────────────────────────────────────────────


def build_filtered_outputs(
    all_hymns: list[dict[str, Any]],
    classifications_by_key: dict[str, dict[str, Any]],
) -> None:
    """Write the filtered index JSON and summary CSV."""
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)

    # ── Build the relevance lookup: hymn_key → {tune_slug → consensus_entry} ──
    relevance_lookup: dict[str, dict[str, dict]] = {}
    for hymn_key, clf in classifications_by_key.items():
        slug_map = {}
        for entry in clf.get("consensus", []):
            slug_map[entry["tune_slug"]] = entry
        relevance_lookup[hymn_key] = slug_map

    # ── Filtered index: keep only relevant tunes ──
    filtered_hymns: list[dict[str, Any]] = []
    for hymn in all_hymns:
        key = hymn["hymn_key"]
        slug_map = relevance_lookup.get(key, {})

        relevant_tunes = []
        for tune in hymn.get("tunes_found", []):
            slug = tune["tune_slug"]
            entry = slug_map.get(slug, {})
            # If not classified yet, keep it (not yet processed)
            if not entry or entry.get("is_relevant", True):
                relevant_tunes.append(tune)

        filtered_hymn = {**hymn, "tunes_found": relevant_tunes}
        filtered_hymns.append(filtered_hymn)

    write_json(FILTERED_INDEX_PATH, filtered_hymns)
    total_kept = sum(len(h["tunes_found"]) for h in filtered_hymns)
    log.info(
        "Wrote filtered index (%d hymns, %d relevant tunes): %s",
        len(filtered_hymns),
        total_kept,
        FILTERED_INDEX_PATH,
    )

    # ── Flat CSV with is_relevant + reasoning columns ──
    rows = []
    for hymn in all_hymns:
        key = hymn["hymn_key"]
        slug_map = relevance_lookup.get(key, {})

        for t in hymn.get("tunes_found", []):
            detail = t.get("detail", {})
            card = t.get("search_card", {})
            slug = t["tune_slug"]
            entry = slug_map.get(slug, {})

            rows.append(
                {
                    "console_display": hymn["console_display"],
                    "full_title": hymn["full_title"],
                    "hymn_key": key,
                    "total_search_results": hymn.get("total_search_results", 0),
                    "tune_title": detail.get(
                        "title", card.get("title", "")
                    ),
                    "tune_slug": slug,
                    "is_relevant": entry.get("is_relevant", True),
                    "vote_count": entry.get("vote_count", ""),
                    "total_runs": entry.get("total_runs", ""),
                    "confidence": entry.get("confidence", ""),
                    "reasoning": entry.get("reasoning", ""),
                    "composer": detail.get("composer", ""),
                    "meter": detail.get("meter", ""),
                    "key": detail.get("key", ""),
                    "num_hymnals": detail.get("num_hymnals", 0),
                    "midi_url": detail.get("midi_url", ""),
                    "recording_url": detail.get("recording_url", ""),
                    "pdf_url": detail.get("pdf_url", ""),
                    "hymnary_url": detail.get(
                        "hymnary_url",
                        f"https://hymnary.org/tune/{slug}",
                    ),
                }
            )

    if rows:
        df = pd.DataFrame(rows)

        # ── Match with Mission Praise Book ──
        if MP_CSV_PATH.exists():
            try:
                try:
                    df_mp = pd.read_csv(MP_CSV_PATH, encoding='utf-8')
                except UnicodeDecodeError:
                    df_mp = pd.read_csv(MP_CSV_PATH, encoding='latin-1')

                if "HymnTuneName" in df_mp.columns:
                    # Create normalized set for O(1) lookup
                    mp_tunes = set(
                        normalize_title(t)
                        for t in df_mp["HymnTuneName"].dropna().astype(str)
                    )
                    
                    # Normalize organ tunes and check
                    df["_norm_title"] = df["tune_title"].apply(normalize_title)
                    df["in_hymn_book"] = df["_norm_title"].apply(lambda x: x in mp_tunes if x else False)
                    df.drop(columns=["_norm_title"], inplace=True)
                    
                    match_count = df["in_hymn_book"].sum()
                    log.info("Matched %d tunes with Mission Praise book", match_count)
                else:
                    log.warning("MP CSV missing 'HymnTuneName' column")
                    df["in_hymn_book"] = False
            except Exception as e:
                log.error("Failed to match MP book: %s", e)
                df["in_hymn_book"] = False
        else:
            df["in_hymn_book"] = False

        df.to_csv(FILTERED_CSV_PATH, index=False)
        relevant_count = df["is_relevant"].sum()
        log.info(
            "Wrote filtered CSV (%d rows, %d relevant, %d excluded): %s",
            len(rows),
            int(relevant_count),
            len(rows) - int(relevant_count),
            FILTERED_CSV_PATH,
        )

        # ── Write formatted Excel file ──
        _write_excel(df, all_hymns, relevance_lookup)


# ──────────────────────────────────────────────
# Excel output
# ──────────────────────────────────────────────


def _write_excel(
    df: pd.DataFrame,
    all_hymns: list[dict[str, Any]],
    relevance_lookup: dict[str, dict[str, dict]],
) -> None:
    """Write a nicely formatted Excel workbook for the user's father.

    Tidy format: one row per hymn–tune pair, all information included.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import shutil

    wb = Workbook()

    # ── Sheet 1: All hymn–tune data (tidy) ──
    ws = wb.active
    ws.title = "Hymn–Tune Data"

    # Column definitions: (header, df_column, width)
    columns = [
        ("Organ Display Name", "console_display", 22),
        ("Hymn Title", "full_title", 35),
        ("Search Results", "total_search_results", 14),
        ("Tune Name", "tune_title", 30),
        ("In Mission Praise?", "in_hymn_book", 18),
        ("Relevant?", "is_relevant", 11),
        ("Confidence", "confidence", 12),
        ("Votes", "vote_count", 8),
        ("AI Reasoning", "reasoning", 50),
        ("Composer", "composer", 28),
        ("Meter", "meter", 20),
        ("Key", "key", 14),
        ("Hymnals Published In", "num_hymnals", 18),
        ("MIDI Link", "midi_url", 38),
        ("Recording Link", "recording_url", 38),
        ("PDF Link", "pdf_url", 38),
        ("Hymnary Page", "hymnary_url", 42),
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6B4C3B", end_color="6B4C3B", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="C9A84C"),
    )

    # Write headers
    for col_idx, (header, _, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    # Data styling
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    relevant_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    excluded_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    link_font = Font(color="4A7C8F", underline="single")

    # Write data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        is_rel = row.get("is_relevant", True)
        row_fill = relevant_fill if is_rel else excluded_fill

        for col_idx, (_, col_key, _) in enumerate(columns, 1):
            value = row.get(col_key, "")

            # Convert booleans to friendly text
            if col_key == "is_relevant":
                value = "Yes" if value else "No"
            elif col_key == "in_hymn_book":
                value = "Yes" if value else ""

            # Handle NaN
            if pd.isna(value):
                value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = wrap_alignment

            # Colour the "Relevant?" column
            if col_key == "is_relevant":
                cell.fill = row_fill
                cell.font = Font(bold=True)

            # Colour the "In Mission Praise?" column
            if col_key == "in_hymn_book" and value == "Yes":
                cell.font = Font(bold=True, color="2E8B57")  # SeaGreen

            # Style URL columns as hyperlinks
            if col_key in ("midi_url", "recording_url", "pdf_url", "hymnary_url") and value:
                cell.font = link_font
                try:
                    cell.hyperlink = str(value)
                except Exception:
                    pass

    # ── Sheet 2: Summary by hymn (one row per hymn) ──
    ws2 = wb.create_sheet("Summary by Hymn")

    summary_cols = [
        ("Organ Display Name", 22),
        ("Hymn Title", 35),
        ("Total Search Results", 18),
        ("Relevant Tunes", 15),
        ("Excluded Tunes", 15),
        ("Top Tune", 30),
        ("Top Tune Composer", 28),
        ("Top Tune Hymnals", 16),
    ]

    for col_idx, (header, width) in enumerate(summary_cols, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(summary_cols))}1"

    # Group by hymn
    hymn_groups = df.groupby("hymn_key", sort=False)
    for row_idx, (hymn_key, group) in enumerate(hymn_groups, 2):
        first = group.iloc[0]
        relevant_group = group[group["is_relevant"] == True]
        excluded_group = group[group["is_relevant"] == False]

        # Pick top tune (relevant, most hymnals)
        if not relevant_group.empty:
            top = relevant_group.sort_values("num_hymnals", ascending=False).iloc[0]
        elif not group.empty:
            top = group.sort_values("num_hymnals", ascending=False).iloc[0]
        else:
            top = pd.Series()

        values = [
            first.get("console_display", ""),
            first.get("full_title", ""),
            first.get("total_search_results", 0),
            len(relevant_group),
            len(excluded_group),
            top.get("tune_title", "") if not top.empty else "",
            top.get("composer", "") if not top.empty else "",
            top.get("num_hymnals", "") if not top.empty else "",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val if not pd.isna(val) else "")
            cell.alignment = wrap_alignment

    # Save
    wb.save(FILTERED_EXCEL_PATH)
    log.info("Wrote Excel workbook: %s", FILTERED_EXCEL_PATH)

    # Copy to site/ directory so Quarto can serve it as a download
    site_dir = PROCESSED_DIR.parent.parent / "site"
    site_dir.mkdir(parents=True, exist_ok=True)
    site_excel = site_dir / "hymn_tune_data.xlsx"
    shutil.copy2(FILTERED_EXCEL_PATH, site_excel)
    log.info("Copied Excel to site/ for download: %s", site_excel)


# ──────────────────────────────────────────────
# Main pipeline
# ──────────────────────────────────────────────


def run(
    limit: int | None = None,
    model: str = "gpt-5.2",
    num_runs: int = 3,
) -> None:
    """Run the LLM classification pipeline with checkpoint/resume."""

    # Load the scraped data
    all_hymns: list[dict[str, Any]] = read_json(HYMN_TUNE_INDEX_PATH) or []
    if not all_hymns:
        log.error("No scraped data found at %s. Run the scraping pipeline first.", HYMN_TUNE_INDEX_PATH)
        sys.exit(1)

    if limit:
        all_hymns_to_process = all_hymns[:limit]
        log.info("Limiting classification to first %d hymns", limit)
    else:
        all_hymns_to_process = all_hymns

    # Load checkpoint + existing classifications
    checkpoint = load_checkpoint(FILTER_CHECKPOINT_PATH)
    completed_keys = set(checkpoint["completed"])

    existing_classifications: dict[str, Any] = read_json(CLASSIFICATIONS_PATH) or {}

    # Set up OpenAI client
    client = get_client()
    log.info("OpenAI client ready (model=%s, num_runs=%d)", model, num_runs)

    total = len(all_hymns_to_process)
    processed = 0
    skipped = 0

    for i, hymn in enumerate(all_hymns_to_process, 1):
        key = hymn["hymn_key"]

        if key in completed_keys:
            log.info("[%d/%d] Skipping (already classified): %s", i, total, hymn["full_title"])
            skipped += 1
            continue

        log.info("━" * 50)
        log.info("[%d/%d] Classifying: %s (%d tunes)", i, total, hymn["full_title"], len(hymn["tunes_found"]))

        try:
            result = classify_hymn(
                client, hymn, model=model, num_runs=num_runs
            )
            existing_classifications[key] = result

            # Mark completed
            if key not in completed_keys:
                checkpoint["completed"].append(key)
                completed_keys.add(key)
            processed += 1

            # Log consensus
            for c in result["consensus"]:
                status = "✓ relevant" if c["is_relevant"] else "✗ excluded"
                log.info(
                    "    %s  %s  (votes: %d/%d, %s)",
                    status,
                    c["tune_slug"],
                    c["vote_count"],
                    c["total_runs"],
                    c["confidence"],
                )

        except KeyboardInterrupt:
            log.warning("\n⚠ Interrupted by user. Progress saved.")
            save_checkpoint(FILTER_CHECKPOINT_PATH, checkpoint)
            write_json(CLASSIFICATIONS_PATH, existing_classifications)
            build_filtered_outputs(all_hymns, existing_classifications)
            sys.exit(0)

        except Exception as exc:
            log.error("[%d/%d] ✗ FAILED: %s — %s", i, total, hymn["full_title"], exc)
            log.error("  %s", traceback.format_exc().splitlines()[-1])
            checkpoint["failed"][key] = str(exc)

        # Save after every hymn
        save_checkpoint(FILTER_CHECKPOINT_PATH, checkpoint)
        write_json(CLASSIFICATIONS_PATH, existing_classifications)

        # Build outputs every 10 hymns
        if processed % 10 == 0 or i == total:
            build_filtered_outputs(all_hymns, existing_classifications)

    # Final output
    build_filtered_outputs(all_hymns, existing_classifications)

    log.info("━" * 50)
    log.info("Classification complete!")
    log.info("  Total hymns:  %d", total)
    log.info("  Classified:   %d", processed)
    log.info("  Skipped:      %d (already done)", skipped)
    log.info("  Failed:       %d", len(checkpoint.get("failed", {})))

    # Summary stats
    total_tunes = 0
    relevant_tunes = 0
    for clf in existing_classifications.values():
        for c in clf.get("consensus", []):
            total_tunes += 1
            if c["is_relevant"]:
                relevant_tunes += 1
    log.info(
        "  Tunes: %d total → %d relevant, %d excluded",
        total_tunes,
        relevant_tunes,
        total_tunes - relevant_tunes,
    )

    if checkpoint.get("failed"):
        for k, err in checkpoint["failed"].items():
            log.info("    ✗ %s: %s", k, err)


# ──────────────────────────────────────────────
# CLI entry point
# ──────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Classify scraped hymn–tune matches using an LLM."
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Classify only the first N hymns (for testing).",
    )
    parser.add_argument(
        "--reset",
        action="store_true",
        help="Clear checkpoint and re-classify everything.",
    )
    parser.add_argument(
        "--model",
        type=str,
        default="gpt-5.2",
        help="OpenAI model to use (default: gpt-5.2).",
    )
    parser.add_argument(
        "--num-runs",
        type=int,
        default=3,
        help="Number of classification runs for majority voting (default: 3).",
    )
    args = parser.parse_args()

    if args.reset:
        FILTER_CHECKPOINT_PATH.unlink(missing_ok=True)
        CLASSIFICATIONS_PATH.unlink(missing_ok=True)
        FILTERED_INDEX_PATH.unlink(missing_ok=True)
        FILTERED_CSV_PATH.unlink(missing_ok=True)
        FILTERED_EXCEL_PATH.unlink(missing_ok=True)
        log.info("Filter checkpoint cleared. Will re-classify all hymns.")

    run(limit=args.limit, model=args.model, num_runs=args.num_runs)


if __name__ == "__main__":
    main()
